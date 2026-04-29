"""
Script para reformatar todas as planilhas do Excel "Planilha Cris.xlsx"
- Mantém as linhas de cabeçalho (header do condomínio)
- Configura colunas: Mês | Ano | Parcela | VALOR DA NOTA | M.O | M.E | VALOR DO INSS | À RECEBER
- ZERA todos os dados das linhas de dados
- Mantém fórmulas automáticas: M.O=D*80%, M.E=D*20%, INSS=E*11%, Receber=D-G
- Aplica grades em todas as células com dados
- Negrito nos cabeçalhos
"""

import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter
import shutil
import os

# Backup do arquivo original
src = "Planilha Cris.xlsx"
bak = "Planilha Cris_BACKUP.xlsx"
if not os.path.exists(bak):
    shutil.copy2(src, bak)
    print(f"Backup criado: {bak}")

print("Carregando planilha... (pode demorar alguns segundos)")
wb = openpyxl.load_workbook(src)
print(f"Total de abas: {len(wb.sheetnames)}")

# ============================================================
# ESTILOS REUTILIZÁVEIS
# ============================================================
thin = Side(style="thin", color="000000")
medium = Side(style="medium", color="000000")

border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_header = Border(left=medium, right=medium, top=medium, bottom=medium)

font_header = Font(bold=True, size=10, name="Calibri")
font_normal = Font(bold=False, size=10, name="Calibri")
font_title   = Font(bold=True, size=11, name="Calibri")

fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # azul claro
fill_data   = PatternFill(fill_type=None)  # sem fill (branco)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left   = Alignment(horizontal="left",   vertical="center")
align_right  = Alignment(horizontal="right",  vertical="center")

# Formato monetário brasileiro
money_fmt = 'R$ #,##0.00'

# ============================================================
# ABAS QUE DEVEM SER IGNORADAS (apenas índices de grupo)
# ============================================================
SKIP_SHEETS = {
    'MENU',
    'CONDOMÍNIO A', 'CONDOMÍNIO B', 'CONDOMÍNIO C',
    'CONDOMÍNIOS D', 'CONDOMÍNIOS E', 'CONDOMÍNIOS F',
    'CONDOMÍNIOS G', 'CONDOMÍNIOS H', 'CONDOMÍNIOS I',
    'CONDOMÍNIOS J', 'CONDOMÍNIOS K', 'CONDOMÍNIOS L',
    'CONDOMÍNIOS M', 'CONDOMÍNIOS N', 'CONDOMÍNIOS O',
    'CONDOMÍNIOS P', 'CONDOMÍNIOS Q', 'CONDOMÍNIOS R',
    'CONDOMÍNIOS S', 'CONDOMÍNIOS T', 'CONDOMÍNIOS U',
    'CONDOMÍNIOS V', 'CONDOMÍNIOS W', 'CONDOMÍNIOS X',
    'CONDOMÍNIOS Y', 'CONDOMÍNIOS Z', 'CONDOMÍNIOS LS',
}

# ============================================================
# COLUNAS DO LAYOUT DA ASSINATURA (lado esquerdo)
# Col A=Mês, B=Ano, C=Parcela, D=Valor Nota, E=M.O, F=M.E, G=INSS, H=À Receber, I=Pago
# ============================================================
HEADERS_ASSINATURA = [
    "Mês", "Ano", "Parcela", "VALOR DA NOTA", "M.O", "M.E", "VALOR DO INSS", "À RECEBER"
]
# Número de linhas de dados (após cabeçalho)
DATA_ROWS = 24  # 2 anos de mensalidades (jan-dez * 2)

def get_row_formula(row_num):
    """Retorna as fórmulas para a linha de dados."""
    d = f"D{row_num}"
    e = f"E{row_num}"
    g = f"G{row_num}"
    return {
        "E": f"=IF({d}=\"\",\"\",{d}*0.8)",     # M.O = Valor da nota * 80%
        "F": f"=IF({d}=\"\",\"\",{d}*0.2)",     # M.E = Valor da nota * 20%
        "G": f"=IF({e}=\"\",\"\",{e}*0.11)",    # INSS = M.O * 11%
        "H": f"=IF({d}=\"\",\"\",{d}-{g})",     # À Receber = Valor da nota - INSS
    }

def apply_cell_style(cell, bold=False, fill=None, border=None, 
                     alignment=None, number_format=None, font_size=10):
    cell.font = Font(bold=bold, size=font_size, name="Calibri")
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format

def format_header_row(ws, row_idx, headers, start_col=1):
    """Escreve e formata a linha de cabeçalho das colunas."""
    for i, header in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=row_idx, column=col, value=header)
        cell.font = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
        cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        cell.border = border_all
        cell.alignment = align_center

def format_data_row(ws, row_idx, formulas, start_col=1, last_col=8):
    """Aplica fórmulas e formatação nas células de dados."""
    col_letters = {
        1: "A", 2: "B", 3: "C", 4: "D",
        5: "E", 6: "F", 7: "G", 8: "H"
    }
    for col in range(start_col, start_col + last_col):
        cell = ws.cell(row=row_idx, column=col)
        letter = col_letters.get(col)
        
        # Zerar dados (A=Mês, B=Ano, C=Parcela, D=Valor da nota)
        if col in [1, 2, 3, 4]:  # A, B, C, D
            cell.value = None
        
        # Aplicar fórmulas
        if letter in formulas:
            cell.value = formulas[letter]
            cell.number_format = money_fmt
        
        # Estilo
        cell.border = border_all
        cell.alignment = align_center
        if col == 1:  # Mês
            cell.alignment = align_left
        cell.font = font_normal

def process_worksheet(ws, sheet_name):
    """Processa uma planilha de dados."""
    
    # -------------------------------------------------------
    # 1. Descobrir as linhas de cabeçalho do condomínio
    #    (geralmente linhas 1-3 com nome, endereço, CNPJ)
    # -------------------------------------------------------
    # Detectar onde começa o header de colunas (procurar por "Mês" ou "MÊS")
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row)):
        for cell in row:
            v = str(cell.value or "").strip().lower()
            if v in ["mês", "mes", "m\u00eas"]:
                header_row = cell.row
                break
        if header_row:
            break
    
    # Se não encontrou linha de header, usa padrão (linha 6)
    if header_row is None:
        # Verificar se há algum conteúdo significativo
        has_content = any(
            cell.value for row in ws.iter_rows(min_row=1, max_row=5)
            for cell in row
        )
        if not has_content and ws.max_row <= 1:
            # Planilha vazia - construir do zero
            header_row = 5
        else:
            header_row = 6

    data_start_row = header_row + 1

    # -------------------------------------------------------
    # 2. Limpar conteúdo abaixo do header de condomínio
    #    (manter linhas 1 a header_row-1, reformar o resto)
    # -------------------------------------------------------
    
    # Limpar linhas de dados existentes (do data_start_row em diante)
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.border = Border()  # remove bordas antigas

    # -------------------------------------------------------
    # 3. Escrever cabeçalho das colunas (linha header_row)
    # -------------------------------------------------------
    format_header_row(ws, header_row, HEADERS_ASSINATURA, start_col=1)

    # -------------------------------------------------------
    # 4. Escrever linhas de dados com fórmulas (DATA_ROWS linhas)
    # -------------------------------------------------------
    for i in range(DATA_ROWS):
        row_idx = data_start_row + i
        formulas = get_row_formula(row_idx)
        format_data_row(ws, row_idx, formulas, start_col=1, last_col=8)

    # -------------------------------------------------------
    # 5. Ajustar larguras das colunas
    # -------------------------------------------------------
    col_widths = {1: 12, 2: 8, 3: 10, 4: 16, 5: 14, 6: 14, 7: 16, 8: 15}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # -------------------------------------------------------
    # 6. Formatar cabeçalho do condomínio (linhas 1 a header_row-1)
    # -------------------------------------------------------
    for row in ws.iter_rows(min_row=1, max_row=header_row - 1):
        for cell in row:
            if cell.value:
                cell.font = Font(bold=True, size=10, name="Calibri")
                cell.alignment = align_left

    # Ajustar altura das linhas
    for i in range(1, header_row):
        ws.row_dimensions[i].height = 16
    ws.row_dimensions[header_row].height = 20
    for i in range(data_start_row, data_start_row + DATA_ROWS):
        ws.row_dimensions[i].height = 18


# ============================================================
# PROCESSAR TODAS AS PLANILHAS
# ============================================================
skip_normalized = {n.upper().strip() for n in SKIP_SHEETS}

processed = 0
skipped = 0
errors = []

for idx, sheet_name in enumerate(wb.sheetnames):
    normalized = sheet_name.upper().strip()
    
    # Pular abas de índice/menu
    if normalized in skip_normalized:
        skipped += 1
        continue
    
    # Pular abas completamente vazias (max_row=1, max_col=1 e sem valor)
    ws = wb[sheet_name]
    
    try:
        process_worksheet(ws, sheet_name)
        processed += 1
        if processed % 50 == 0:
            print(f"  Processadas: {processed} planilhas...")
    except Exception as e:
        errors.append((sheet_name, str(e)))
        print(f"  ERRO na aba '{sheet_name}': {e}")

print(f"\nResumo:")
print(f"  Processadas: {processed}")
print(f"  Ignoradas:   {skipped}")
print(f"  Erros:       {len(errors)}")
if errors:
    for name, err in errors[:10]:
        print(f"    - {name}: {err}")

# ============================================================
# SALVAR
# ============================================================
print("\nSalvando arquivo... (pode demorar)")
wb.save(src)
print(f"Arquivo salvo: {src}")
print("Concluído!")
