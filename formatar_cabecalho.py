"""
Script para formatar o cabeçalho do condomínio (linhas 1-5) em todas as planilhas:
- Adiciona grades/bordas em todas as células do cabeçalho
- Coloca negrito em todo o texto do cabeçalho
- Garante que as células estão desbloqueadas (editáveis)
- Adiciona cor de fundo levemente azulada para destacar
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter

print("Carregando planilha...")
wb = openpyxl.load_workbook("Planilha Cris.xlsx")
print(f"Total de abas: {len(wb.sheetnames)}")

# ============================================================
# ESTILOS
# ============================================================
thin  = Side(style="thin",   color="4472C4")   # azul médio para borda interna
medium = Side(style="medium", color="1F3864")   # azul escuro para borda externa

# Borda interna das células de cabeçalho
border_inner = Border(left=thin, right=thin, top=thin, bottom=thin)
# Borda grossa para a linha de separação (linha 5 = última do header info)
border_bottom_thick = Border(
    left=thin, right=thin, top=thin,
    bottom=Side(style="medium", color="1F3864")
)

# Fundo levemente azulado para linhas 1-3 (info do condomínio)
fill_info  = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
# Fundo um pouco diferente para linha 4-5 (Ano, Vencimento)
fill_meta  = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
# Fundo azul escuro para linha 6 (cabeçalho de colunas) - já feito antes
fill_col_header = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

font_info  = Font(bold=True, size=10, name="Calibri", color="1F3864")
font_meta  = Font(bold=True, size=10, name="Calibri", color="1F3864")
font_white = Font(bold=True, size=10, name="Calibri", color="FFFFFF")

align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Proteção: célula desbloqueada (editável mesmo se a planilha for protegida)
unlocked = Protection(locked=False)

# Abas a ignorar
SKIP_PREFIXES = ("MENU", "CONDOM")

# ============================================================
# DESCOBRIR QUANTAS COLUNAS HÁ EM CADA PLANILHA
# ============================================================
def get_last_used_col(ws):
    """Retorna o índice da última coluna com algum valor."""
    max_col = 0
    for row in ws.iter_rows(max_row=8):
        for cell in row:
            if cell.value is not None and cell.column > max_col:
                max_col = cell.column
    return max(max_col, 8)  # pelo menos 8 colunas (A-H)


def format_header(ws):
    """Formata as linhas de cabeçalho (1-5) da planilha."""
    
    last_col = get_last_used_col(ws)
    
    # Encontrar onde fica a linha do cabeçalho de colunas (Mês, Ano, Parcela...)
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            v = str(cell.value or "").strip().lower()
            if v in ["mês", "mes", "m\u00eas"]:
                header_row = cell.row
                break
        if header_row:
            break
    if header_row is None:
        header_row = 6

    info_rows = list(range(1, header_row))  # linhas 1 a 5

    # ----------------------------------------------------------
    # Formatar linhas de info do condomínio (1 até header_row-1)
    # ----------------------------------------------------------
    for row_idx in info_rows:
        is_last_info_row = (row_idx == header_row - 1)
        brd = border_bottom_thick if is_last_info_row else border_inner
        
        # Escolher fill: linhas 1-3 = azul claro, 4-5 = azul médio
        if row_idx <= 3:
            fill = fill_info
        else:
            fill = fill_meta

        for col in range(1, last_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            # Aplicar borda
            cell.border = brd
            # Aplicar fill (fundo colorido)
            cell.fill = fill
            # Negrito + cor azul escura
            cell.font = font_info
            # Alinhamento
            cell.alignment = align_left
            # Desbloquear para edição
            cell.protection = unlocked
            # Altura das linhas
            ws.row_dimensions[row_idx].height = 18

    # ----------------------------------------------------------
    # Reformatar linha de cabeçalho de colunas (header_row)
    # ----------------------------------------------------------
    for col in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value:  # só formata se tiver valor
            cell.fill = fill_col_header
            cell.font = font_white
            cell.border = Border(
                left=Side(style="medium", color="FFFFFF"),
                right=Side(style="medium", color="FFFFFF"),
                top=Side(style="medium", color="1F3864"),
                bottom=Side(style="medium", color="1F3864"),
            )
            cell.alignment = align_center
            cell.protection = unlocked
        ws.row_dimensions[header_row].height = 22


# ============================================================
# PROCESSAR TODAS AS PLANILHAS
# ============================================================
processed = 0
skipped = 0
errors = []

for sheet_name in wb.sheetnames:
    norm = sheet_name.upper().strip()
    # Pular MENU e índices de condomínio
    if any(norm.startswith(p) for p in SKIP_PREFIXES):
        skipped += 1
        continue

    ws = wb[sheet_name]
    try:
        format_header(ws)
        processed += 1
        if processed % 50 == 0:
            print(f"  Formatadas: {processed} planilhas...")
    except Exception as e:
        errors.append((sheet_name, str(e)))

print(f"\nResumo:")
print(f"  Formatadas: {processed}")
print(f"  Ignoradas:  {skipped}")
print(f"  Erros:      {len(errors)}")
if errors:
    for name, err in errors[:5]:
        print(f"    - {name}: {err}")

print("\nSalvando... (pode levar alguns segundos)")
wb.save("Planilha Cris.xlsx")
print("Salvo! Concluido.")
